import logging
import ast
import time
import os
import io
from typing import Union

from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import (
    QDesktopWidget, QGridLayout, QGroupBox,
    QLabel, QLineEdit, QMainWindow,QPushButton,
    QRadioButton, QWidget,
)

import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
#  from matplotlib.backends.backend_qt5agg import NavigationToolbar2QT as NavigationToolBar

import serial

import numpy as np

import pandas as pd

from numba import jit

from openpyxl import Workbook

from . import config
from .contrib import list_com_port, ErrorDialog, SettingWindow

class WindowApp(QMainWindow):
    def __init__(self, parent=None) -> None:
        super().__init__(parent)

        self.setWindowTitle("FMWC Radar Launcher")
        self.resize(640, 480)

        self.start_get_data = True

        # Move window to the center of the screen
        fg = self.frameGeometry()
        cp = QDesktopWidget().availableGeometry().center()
        fg.moveCenter(cp)
        self.move(fg.topLeft())

        self.grid = QGridLayout()

        widget = QWidget()
        widget.setLayout(self.grid)

        self.grid.addWidget(self._top_button_group(), 0, 0)
        self.grid.addWidget(self._profile_group(), 1, 0)
        self.grid.addWidget(self._save_data_group(), 2, 0)

        self.grid.addWidget(self._respiro_graph(), 0, 1,)
        self.grid.addWidget(self._twr_graps(), 1, 1, 2, 1)

        self.grid.setColumnStretch(1, 3)

        self.setCentralWidget(widget)
        self.showMaximized()

    def closeEvent(self, evnt) -> None:

        try:
            self.serial.close()
        except AttributeError:
            pass

        super(WindowApp, self).closeEvent(evnt)

    def _save_respiro_data(self) -> None:
        #  dest_path = os.path.expanduser(f"~\\Documents\\Data Respiro_{self.file_name.text()}.xlsx")
        dest_path = os.path.expanduser(f"~\\Documents\\Data_{self.file_name.text()}.csv")

        #  wb = Workbook()
        #  ws = wb.active

        #  ws['A1'] = f'Nama Pasien: {self.input_name.text()}'
        #  ws['A2'] = f'Usia Pasien: {self.input_age.text()}'
        #  ws['A3'] = f'Jenis Kelamin: {self.input_sex.text()}'

        #  wb.save(dest_path)

        try:
            if self.radio_resp.isChecked():
                data = pd.DataFrame({
                    "Respiro": self.respiro_out,
                })
            else:
                data = pd.DataFrame({
                    "FFt Magnitude": self.twr_out,
                })
            data.to_csv(dest_path)
        except AttributeError as e:
            error_dialog = ErrorDialog(f"Tidak ada data yang ditangkap: {e}", "Data Error", self)
            error_dialog.exec_()

    def _profile_group(self):
        group_box = QGroupBox("Identitas Pasien")
        layout = QGridLayout()

        label_name = QLabel("Pasien Atas Nama: ")
        self.input_name = QLineEdit()

        label_age = QLabel("Usia Pasien: ")
        self.input_age = QLineEdit()

        label_sex = QLabel("Jenis Kelamin: ")
        self.input_sex = QLineEdit()

        layout.addWidget(label_name, 0, 0)
        layout.addWidget(self.input_name, 0, 1)

        layout.addWidget(label_age, 1, 0)
        layout.addWidget(self.input_age, 1, 1)

        layout.addWidget(label_sex, 2, 0)
        layout.addWidget(self.input_sex, 2, 1)

        group_box.setLayout(layout)
        return group_box

    def _top_button_group(self) -> QWidget:
        group_box = QGroupBox("Menu Program")

        bsetting = QPushButton("Setting")
        bstart = QPushButton("Mulai Scan")
        bstop = QPushButton("Stop Scan")

        self.distance_label = QLineEdit("Jarak: 0")

        bstart.clicked.connect(self._start_process)
        bstop.clicked.connect(self._stop_get_data)
        bsetting.clicked.connect(self._launch_setting)

        h_box = QGridLayout()
        h_box.setAlignment(Qt.AlignTop)

        h_box.addWidget(bstart, 0, 0)
        h_box.addWidget(bstop, 0, 1)
        h_box.addWidget(bsetting, 1, 0, 1, 2)
        h_box.addWidget(self.distance_label, 2, 0, 1, 2)
        h_box.addWidget(QLabel(
            "Mohon untuk mengcek serial port yang tersedia, setelah serial port sesuai, aplikasi siap dijalankan."),
            3, 0, 1, 2)

        group_box.setLayout(h_box)
        return group_box

    def _save_data_group(self):
        group_box = QGroupBox("Simpan Data Respiro")

        bsave = QPushButton("Simpan Data")
        
        file_name_label = QLabel("Masukan Nama File")
        self.file_name = QLineEdit()

        radio_label = QLabel("Silahkan pilih data yang akan disimpan", self)
        self.radio_fft = QRadioButton("Data Magnitude &FFT", self)
        self.radio_resp = QRadioButton("Data &Respiro", self)
        self.radio_resp.toggle()

        layout = QGridLayout()
        layout.setAlignment(Qt.AlignTop)

        layout.addWidget(file_name_label, 0, 0)
        layout.addWidget(self.file_name, 0, 1)

        layout.addWidget(radio_label, 1, 0, 1, 2)
        layout.addWidget(self.radio_fft, 2, 0)
        layout.addWidget(self.radio_resp, 2, 1)

        layout.addWidget(bsave, 3, 0, 1, 2)

        bsave.clicked.connect(self._save_respiro_data)

        group_box.setLayout(layout)
        return group_box

    def _start_process(self):
        self._set_serial_port()
        try:

            if not self.start_get_data:
                self.start_get_data = True

            self._get_data_respiro()
        except AttributeError:
            pass

    def _launch_setting(self):
        dialog = SettingWindow(self)
        dialog.exec_()

    def _set_serial_port(self):
        try:
            self.serial = serial.Serial(
                port=config.serial_port,
                baudrate=9600,
                parity=serial.PARITY_ODD,
                stopbits=serial.STOPBITS_TWO,
                bytesize=serial.SEVENBITS
            )
        except serial.SerialException as e:
            error_window = ErrorDialog(f"Error Serial Port: {e}", "Serial Port Not Found", self)
            error_window.exec_()

    def _twr_graps(self):
        group_box = QGroupBox("TWR Graph")

        self.figure_twr = plt.figure()
        self.canvas_twr = FigureCanvas(self.figure_twr)

        grid = QGridLayout()
        grid.addWidget(self.canvas_twr, 1, 0)

        group_box.setLayout(grid)

        return group_box

    def _respiro_graph(self):
        group_box = QGroupBox("Respiro Meter")

        self.figure_resp = plt.figure()
        self.canvas_resp = FigureCanvas(self.figure_resp)

        grid = QGridLayout()
        grid.addWidget(self.canvas_resp, 1, 0)

        group_box.setLayout(grid)

        return group_box

    def _max_index_value(self, ls) -> Union[dict, None]:
        """
        return maximum value and it's index from a list
        """

        maxval = max(ls)
        idx = ls.index(maxval)

        return {"val": maxval, "index": idx}

    @staticmethod
    @jit(nopython=True)
    def _process_data_respiro(y_vec, yo_vec):
        # Get ready for a loooong calculation
        return yo_vec[-1] * 0.0048 + yo_vec[-2] * 0.0195 + yo_vec[-3] * 0.0289 + yo_vec[-4] * 0.0193 + yo_vec[-5] * 0.0048 - y_vec[-1] - y_vec[-2] * -2.3695 - y_vec[-3] * 2.3140 - y_vec[-4] * -1.0547 - y_vec[-5] * 0.1874

    def _get_data_respiro(self):
        if not self.serial.isOpen():
            logging.info("[INFO] Open the serial port...")
            self.serial.open()

        # instead of ax.hold(False)
        self.figure_resp.clear()
        self.figure_twr.clear()

        # create an axis
        ax_reps = self.figure_resp.add_subplot(111)
        ax_twr = self.figure_twr.add_subplot(111)

        # Do some clean up, just to make sure
        self.serial.flushInput()
        self.serial.flushOutput()

        self.serial.write(str.encode("oF"))
        time.sleep(0.5)
        self.serial.write(str.encode("oP"))

        y_vec = np.linspace(0, 0, 512)[:-1]
        yo_vec = np.linspace(0, 1, 512)[:-1]
        x_vec = np.linspace(0, 1, 512)[:-1]

        self.respiro_out = []
        self.twr_out = []

        while self.start_get_data:

            if not self.serial.isOpen():
                break

            try:
                distance = ast.literal_eval(self.serial.readline().decode("utf-8"))
            except (SyntaxError, UnicodeDecodeError):
                continue

            if isinstance(distance, tuple):
                #  print("Jarak:", distance[1], end="\r")
                self.distance_label.setText(f"Jarak: {distance[1]}")
                #  pass
            elif isinstance(distance, dict):
                fft_phase = distance.get("Phase")
                self.fft_mag = distance.get("FFT") 

                if fft_phase:
                    yo_vec[-1] = float(fft_phase[2]) * 57.29
                    y_vec[-1] = self._process_data_respiro(y_vec, yo_vec)

                    self.respiro_out.append(fft_phase[:512])
                    #  print(y_vec[-1], end="\r")

                    #  refresh and plot the data
                    ax_reps.clear()

                    line1, = ax_reps.plot(x_vec[462:], y_vec[462:], '-o', alpha=0.8)
                    line1.set_ydata(y_vec[462:])

                    if np.min(y_vec) <= line1.axes.get_ylim()[0] or np.max(y_vec) >= line1.axes.get_ylim()[1]:
                        plt.ylim([np.min(y_vec) - np.std(y_vec), np.max(y_vec) + np.std(y_vec)])

                    # refresh canvas
                    self.canvas_resp.draw_idle()
                    self.canvas_resp.flush_events()

                    y_vec = np.append(y_vec[1:], 0.0)
                    #  np.stack([0], y_vec)
                    #  print(len(y_vec), end="\r")

                elif self.fft_mag:
                    self.twr_out.append(self.fft_mag[:512])
                    ax_twr.clear()
                    ax_twr.plot(self.fft_mag[:100])

                    self.canvas_twr.draw_idle()
                    self.canvas_twr.flush_events()

        #  print(self.respiro_out)
        self.serial.close()

    def _refresh_ports_list(self) -> None:
        self.port_list.setPlainText(list_com_port())

    def _stop_get_data(self) -> None:
        self.start_get_data = False

